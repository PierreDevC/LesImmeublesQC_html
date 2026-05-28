import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_listings_excel():
    # Create workbook and select active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active Listings"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Define columns with Website Link as the 2nd column (Column B)
    headers = [
        "Centris No.",
        "Website Link",
        "Address",
        "Rent (per month)",
        "Type",
        "Bedrooms",
        "Bathrooms",
        "Total Rooms",
        "Gross Area",
        "Year Built",
        "Furnished",
        "Utilities Included",
        "Appliances",
        "Pets",
        "Smoking",
        "Walk Score",
        "Listing Agency"
    ]
    
    # Define row data (with 3 separate entries for Aubry to keep pricing numeric)
    data = [
        {
            "Centris No.": 24098291,
            "Website Link": "https://www.centris.ca/en/condos-apartments~for-rent~montreal-ville-marie/24098291",
            "Address": "1935, Rue Tupper, apt. 16, Montréal (Ville-Marie)",
            "Rent (per month)": 1558,
            "Type": "Condo / Apartment",
            "Bedrooms": 1,
            "Bathrooms": 1,
            "Total Rooms": 4,
            "Gross Area": "400 sqft",
            "Year Built": 1928,
            "Furnished": "Yes (Fully Furnished)",
            "Utilities Included": "High-speed Internet",
            "Appliances": "Fridge, stove, microwave, air conditioning",
            "Pets": "Allowed (with conditions)",
            "Smoking": "Not allowed",
            "Walk Score": 100,
            "Listing Agency": "Keller Williams Prestige"
        },
        {
            "Centris No.": 22663570,
            "Website Link": "https://www.realtor.ca/immobilier/29322009/1935-rue-tupper-17-montreal-ville-marie-centre-ouest",
            "Address": "1935 Rue Tupper, app. 17, Montréal (Ville-Marie), QC, H3H 1N6",
            "Rent (per month)": 1375,
            "Type": "Condo / Apartment",
            "Bedrooms": 1,
            "Bathrooms": 1,
            "Total Rooms": 4,
            "Gross Area": "500 sqft",
            "Year Built": 1928,
            "Furnished": "No (Semi-furnished)",
            "Utilities Included": "High-speed Internet",
            "Appliances": "Fridge, stove, microwave, air conditioning",
            "Pets": "Allowed (with conditions)",
            "Smoking": "Not allowed",
            "Walk Score": 100,
            "Listing Agency": "Keller Williams Prestige"
        },
        {
            "Centris No.": "N/A",
            "Website Link": "https://www.zumper.com/address/2930-rue-aubry-montreal-qc-h1l-4g9-can",
            "Address": "2930 Rue Aubry (Top Floor), Montréal (Tétreaultville), QC, H1L 4G9",
            "Rent (per month)": 1675,
            "Type": "Condo / Apartment",
            "Bedrooms": 2,
            "Bathrooms": 1,
            "Total Rooms": 4,
            "Gross Area": "N/A",
            "Year Built": "N/A",
            "Furnished": "No (Semi-furnished)",
            "Utilities Included": "Wireless Internet",
            "Appliances": "Fridge, stove, washer, dryer, air conditioning",
            "Pets": "Not allowed",
            "Smoking": "Not allowed",
            "Walk Score": "N/A",
            "Listing Agency": "Joshua (Rental Agent)"
        },
        {
            "Centris No.": "N/A",
            "Website Link": "https://www.zumper.com/address/2930-rue-aubry-montreal-qc-h1l-4g9-can",
            "Address": "2930 Rue Aubry (Ground / 2nd Floor), Montréal (Tétreaultville), QC, H1L 4G9",
            "Rent (per month)": 1650,
            "Type": "Condo / Apartment",
            "Bedrooms": 2,
            "Bathrooms": 1,
            "Total Rooms": 4,
            "Gross Area": "N/A",
            "Year Built": "N/A",
            "Furnished": "No (Semi-furnished)",
            "Utilities Included": "Wireless Internet",
            "Appliances": "Fridge, stove, washer, dryer, air conditioning",
            "Pets": "Not allowed",
            "Smoking": "Not allowed",
            "Walk Score": "N/A",
            "Listing Agency": "Joshua (Rental Agent)"
        },
        {
            "Centris No.": "N/A",
            "Website Link": "https://www.zumper.com/address/2930-rue-aubry-montreal-qc-h1l-4g9-can",
            "Address": "2930 Rue Aubry (Semi-Basement), Montréal (Tétreaultville), QC, H1L 4G9",
            "Rent (per month)": 1550,
            "Type": "Condo / Apartment",
            "Bedrooms": 2,
            "Bathrooms": 1,
            "Total Rooms": 4,
            "Gross Area": "N/A",
            "Year Built": "N/A",
            "Furnished": "No (Semi-furnished)",
            "Utilities Included": "Wireless Internet",
            "Appliances": "Fridge, stove, washer, dryer, air conditioning",
            "Pets": "Not allowed",
            "Smoking": "Not allowed",
            "Walk Score": "N/A",
            "Listing Agency": "Joshua (Rental Agent)"
        }
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        
    # Styles
    font_family = "Segoe UI"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Sleek Dark Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Apply styling to headers
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Write data and apply formatting
    data_font = Font(name=font_family, size=10, color="000000")
    link_font = Font(name=font_family, size=10, color="0563C1", underline="single")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    for row_idx, row_data in enumerate(data, 2):
        ws.row_dimensions[row_idx].height = 22
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = row_data.get(header)
            
            # Format and align based on column type
            if header == "Website Link":
                cell.value = val
                cell.hyperlink = val
                cell.font = link_font
                cell.alignment = left_align
            else:
                cell.value = val
                cell.font = data_font
                
                if header in ["Centris No.", "Bedrooms", "Bathrooms", "Total Rooms", "Year Built", "Walk Score", "Furnished", "Pets", "Smoking"]:
                    cell.alignment = center_align
                elif header in ["Rent (per month)"]:
                    if isinstance(val, (int, float)):
                        cell.alignment = right_align
                        cell.number_format = '$#,##0'
                    else:
                        cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
            cell.border = thin_border
            
    # Auto-adjust column widths with safety margin
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            # Keep the hyperlink column compact
            if cell.row > 1 and cell.column == 2:  # Website Link column
                val_str = "View Zumper Listing" # Length proxy
            else:
                val_str = str(cell.value or '')
            
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        # Set a reasonable min/max width
        width = max(max_len + 4, 12)
        if col_letter == 'C':  # Address column
            width = min(width, 50)  # Cap address width to prevent super wide col
            # Enable wrap text on address data rows
            for cell in col:
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif col_letter == 'B':  # Website Link column
            width = 25  # Set a nice standard width for the hyperlink column
            
        ws.column_dimensions[col_letter].width = width

    # Save to workspace /docs folder
    output_path = "/Volumes/crucial/Shortkut/LesImmeublesQC_html/docs/active_listings.xlsx"
    wb.save(output_path)
    print(f"Excel file successfully updated at: {output_path}")

if __name__ == "__main__":
    create_listings_excel()
